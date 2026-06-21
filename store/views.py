import json
from decimal import Decimal
from io import BytesIO
from textwrap import wrap
from urllib.parse import quote

from django.contrib.admin.views.decorators import staff_member_required


from django.contrib import messages
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.db.models import Count, Prefetch, Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from PIL import Image, ImageDraw, ImageFont

from marque.models import Marque

from .models import (
    Categorie,
    SousCategorie,
    Product,
    Cart,
    CartItem,
    OffreSpeciale,
    PromotionItem,
    AvisClient,
    CustomerOrder,
    CustomerOrderItem,
    StockMovement,
)
# ============================================================
# PAGINATION HELPER
# ============================================================

PRODUCTS_PER_PAGE = 16
OFFERS_PER_PAGE = 16
BRANDS_PER_PAGE = 16

def paginate_queryset(request, queryset, per_page=12):
    paginator = Paginator(queryset, per_page)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop("page", None)
    pagination_query = query_params.urlencode()

    return page_obj, pagination_query


# ============================================================
# CART HELPERS
# ============================================================

def merge_cart_items(source_cart, target_cart):
    if not source_cart or not target_cart or source_cart.id == target_cart.id:
        return target_cart

    for item in source_cart.store_items.all():
        existing_item = None

        if item.product:
            existing_item = CartItem.objects.filter(
                cart=target_cart,
                product=item.product,
                offre__isnull=True
            ).first()

        elif item.offre:
            existing_item = CartItem.objects.filter(
                cart=target_cart,
                offre=item.offre,
                product__isnull=True
            ).first()

        if existing_item:
            existing_item.quantity += item.quantity
            existing_item.save(update_fields=["quantity"])
            item.delete()
        else:
            item.cart = target_cart
            item.save(update_fields=["cart"])

    source_cart.delete()
    return target_cart


def get_or_create_cart(request):
    if not request.session.session_key:
        request.session.create()

    session_key = request.session.session_key

    if request.user.is_authenticated:
        user_cart, _ = Cart.objects.get_or_create(
            user=request.user,
            is_ordered=False,
            defaults={
                "session_key": session_key,
            }
        )

        guest_cart = Cart.objects.filter(
            user__isnull=True,
            session_key=session_key,
            is_ordered=False
        ).first()

        if guest_cart and guest_cart.id != user_cart.id:
            user_cart = merge_cart_items(guest_cart, user_cart)

        return user_cart

    guest_cart = Cart.objects.filter(
        user__isnull=True,
        session_key=session_key,
        is_ordered=False
    ).first()

    if not guest_cart:
        guest_cart = Cart.objects.create(
            user=None,
            session_key=session_key,
            is_ordered=False
        )

    return guest_cart


def calculate_cart_totals(cart_obj):
    total = Decimal("0.00")

    for item in cart_obj.store_items.all():
        total += item.get_price()

    shipping_cost = Decimal("0.00")

    if total > 0 and total < Decimal("500.00"):
        shipping_cost = Decimal("30.00")

    total_with_shipping = total + shipping_cost

    return total, shipping_cost, total_with_shipping


# ============================================================
# HOME / PRODUCTS
# ============================================================

def index(request):
    now = timezone.now()

    home_products = (
        Product.objects
        .filter(active=True, show_on_home=True)
        .select_related("marque", "categorie", "sous_categorie")
        .order_by("home_order", "-id")[:12]
    )

    products = (
        Product.objects
        .filter(active=True)
        .select_related("marque", "categorie", "sous_categorie")
        .order_by("-id")[:8]
    )

    promo_items = (
        PromotionItem.objects
        .filter(
            active=True,
            product__active=True,
            promotion__active=True,
            promotion__show_home=True,
            promotion__start_date__lte=now,
            promotion__end_date__gte=now,
        )
        .select_related(
            "promotion",
            "product",
            "product__marque",
            "product__categorie",
            "product__sous_categorie",
        )
        .order_by("-created_at")[:8]
    )

    avis = (
        AvisClient.objects
        .select_related("produit", "user")
        .order_by("-date_ajout")[:5]
    )

    offres = (
        OffreSpeciale.objects
        .filter(
            Q(date_expiration__isnull=True) |
            Q(date_expiration__gt=now)
        )
        .order_by("-id")
    )

    categories = (
        Categorie.objects
        .filter(active=True)
        .annotate(
            sub_count=Count(
                "subcategories",
                filter=Q(subcategories__active=True),
                distinct=True
            ),
            product_count=Count(
                "products",
                filter=Q(products__active=True),
                distinct=True
            ),
        )
        .order_by("ordre", "nom")
    )

    home_categories = []

    for cat in categories:
        first_sub = (
            SousCategorie.objects
            .filter(parent=cat, active=True)
            .order_by("ordre", "nom")
            .first()
        )

        home_categories.append({
            "slug": cat.slug,
            "label": cat.nom,
            "image": cat.image if cat.image else (
                first_sub.image if first_sub and first_sub.image else None
            ),
            "sub_count": cat.sub_count,
            "product_count": cat.product_count,
        })

    return render(request, "store/index.html", {
        "home_products": home_products,
        "products": products,
        "promo_items": promo_items,
        "avis": avis,
        "offres": offres,
        "home_categories": home_categories,
    })


def product_detail(request, slug):
    product = get_object_or_404(
        Product.objects.select_related("marque", "categorie", "sous_categorie"),
        slug=slug,
        active=True
    )

    related_products = (
        Product.objects
        .filter(categorie=product.categorie, active=True)
        .exclude(id=product.id)
        .select_related("marque", "categorie", "sous_categorie")[:4]
    )

    return render(request, "store/product_detail.html", {
        "product": product,
        "related_products": related_products,
    })


def produits_par_categorie(request, slug):
    categorie = get_object_or_404(Categorie, slug=slug, active=True)

    sous_categories = (
        SousCategorie.objects
        .filter(parent=categorie, active=True)
        .annotate(
            nbr_produits=Count(
                "products",
                filter=Q(products__active=True),
                distinct=True
            )
        )
        .order_by("ordre", "nom")
    )

    produits_queryset = (
        Product.objects
        .filter(categorie=categorie, active=True)
        .select_related("marque", "categorie", "sous_categorie")
        .order_by("-id")
    )

    sub_slug = request.GET.get("sub")
    selected_sous_categorie = None

    if sub_slug:
        selected_sous_categorie = get_object_or_404(
            SousCategorie,
            slug=sub_slug,
            parent=categorie,
            active=True
        )
        produits_queryset = produits_queryset.filter(
            sous_categorie=selected_sous_categorie
        )

    produits, pagination_query = paginate_queryset(
        request,
        produits_queryset,
        PRODUCTS_PER_PAGE
    )

    return render(request, "store/produits.html", {
        "produits": produits,
        "page_obj": produits,
        "pagination_query": pagination_query,
        "categorie": categorie,
        "categorie_code": categorie.slug,
        "sous_categories": sous_categories,
        "selected_sous_categorie": selected_sous_categorie,
    })


def products_by_subcategory(request, slug):
    sous_categorie = get_object_or_404(
        SousCategorie.objects.select_related("parent"),
        slug=slug,
        active=True
    )

    url = reverse(
        "store:produits_par_categorie",
        args=[sous_categorie.parent.slug]
    )

    return redirect(f"{url}?sub={sous_categorie.slug}")


# ============================================================
# CART
# ============================================================

def cart(request):
    cart_obj = get_or_create_cart(request)

    cart_items = (
        cart_obj.store_items
        .select_related("product", "product__marque", "offre")
        .all()
    )

    total, shipping_cost, total_with_shipping = calculate_cart_totals(cart_obj)

    return render(request, "store/cart.html", {
        "cart": cart_obj,
        "cart_items": cart_items,
        "total": total,
        "shipping_cost": shipping_cost,
        "total_with_shipping": total_with_shipping,
    })


def cart_view(request):
    return cart(request)


def add_to_cart(request, slug):
    product = get_object_or_404(Product, slug=slug, active=True)

    if product.stock <= 0:
        messages.warning(request, "Ce produit est en rupture de stock.")
        return redirect("store:product_detail", slug=slug)

    cart_obj = get_or_create_cart(request)

    if request.method == "POST":
        try:
            quantity = int(request.POST.get("quantity", 1))
        except (ValueError, TypeError):
            quantity = 1
    else:
        quantity = 1

    if quantity < 1:
        quantity = 1

    item, created = CartItem.objects.get_or_create(
        cart=cart_obj,
        product=product,
        offre=None
    )

    old_quantity = 0 if created else item.quantity
    new_quantity = old_quantity + quantity

    if new_quantity > product.stock:
        item.quantity = product.stock
        messages.warning(
            request,
            f"Stock disponible pour {product.nom} : {product.stock}. La quantité a été ajustée."
        )
    else:
        item.quantity = new_quantity
        messages.success(request, "Produit ajouté au panier.")

    item.save(update_fields=["quantity"])

    return redirect("store:product_detail", slug=slug)


def add_offer_to_cart(request, slug):
    offre = get_object_or_404(OffreSpeciale, slug=slug)
    cart_obj = get_or_create_cart(request)

    if request.method == "POST":
        try:
            quantity = int(request.POST.get("quantity", 1))
        except (ValueError, TypeError):
            quantity = 1
    else:
        quantity = 1

    if quantity < 1:
        quantity = 1

    item, created = CartItem.objects.get_or_create(
        cart=cart_obj,
        offre=offre,
        product=None
    )

    if created:
        item.quantity = quantity
    else:
        item.quantity += quantity

    item.save(update_fields=["quantity"])

    messages.success(request, "Offre ajoutée au panier.")
    return redirect("store:detail_offre", slug=slug)


@require_POST
def delete_cart(request):
    cart_obj = get_or_create_cart(request)
    cart_obj.store_items.all().delete()

    messages.success(request, "Votre panier a été supprimé.")
    return redirect("store:index")


@require_POST
def remove_item(request, item_id):
    cart_obj = get_or_create_cart(request)

    item = get_object_or_404(
        CartItem,
        id=item_id,
        cart=cart_obj
    )

    item.delete()
    messages.success(request, "Produit supprimé du panier.")
    return redirect("store:cart")


@require_POST
def remove_from_cart(request, item_id):
    return remove_item(request, item_id)


@require_POST
def update_cart_item(request, item_id):
    return update_quantity(request, item_id)


@require_POST
def update_quantity(request, item_id):
    cart_obj = get_or_create_cart(request)

    cart_item = get_object_or_404(
        CartItem,
        id=item_id,
        cart=cart_obj
    )

    try:
        quantity = int(request.POST.get("quantity", 1))
    except (ValueError, TypeError):
        quantity = 1

    if quantity < 1:
        quantity = 1

    if cart_item.product:
        if cart_item.product.stock <= 0:
            messages.warning(request, f"{cart_item.product.nom} est en rupture de stock.")
            cart_item.delete()
            return redirect("store:cart")

        if quantity > cart_item.product.stock:
            quantity = cart_item.product.stock
            messages.warning(
                request,
                f"Stock disponible pour {cart_item.product.nom} : {cart_item.product.stock}. La quantité a été ajustée."
            )

    cart_item.quantity = quantity
    cart_item.save(update_fields=["quantity"])

    messages.success(request, "Quantité mise à jour.")
    return redirect("store:cart")


@require_POST
def update_quantity_ajax(request, item_id):
    try:
        quantity = 1

        if request.body:
            try:
                data = json.loads(request.body.decode("utf-8"))
                quantity = int(data.get("quantity", 1))
            except json.JSONDecodeError:
                quantity = int(request.POST.get("quantity", 1))
        else:
            quantity = int(request.POST.get("quantity", 1))

        if quantity < 1:
            quantity = 1

        cart_obj = get_or_create_cart(request)
        item = get_object_or_404(CartItem, id=item_id, cart=cart_obj)

        if item.product and quantity > item.product.stock:
            quantity = item.product.stock

        item.quantity = quantity
        item.save(update_fields=["quantity"])

        total, shipping_cost, total_with_shipping = calculate_cart_totals(cart_obj)

        return JsonResponse({
            "success": True,
            "new_quantity": item.quantity,
            "item_total": f"{item.get_price():.2f}",
            "cart_total": f"{total:.2f}",
            "shipping_cost": f"{shipping_cost:.2f}",
            "total_with_shipping": f"{total_with_shipping:.2f}",
        })

    except Exception as e:
        return JsonResponse({
            "success": False,
            "error": str(e),
        }, status=400)


# ============================================================
# CHECKOUT / WHATSAPP ORDER
# ============================================================

def checkout(request):
    cart_obj = get_or_create_cart(request)

    if not cart_obj.store_items.exists():
        messages.warning(request, "Votre panier est vide.")
        return redirect("store:cart")

    cart_items = (
        cart_obj.store_items
        .select_related("product", "offre")
        .all()
    )

    total, shipping_cost, total_with_shipping = calculate_cart_totals(cart_obj)

    if request.method == "POST":
        if request.user.is_authenticated:
            full_name = (
                request.POST.get("full_name")
                or request.user.get_full_name()
                or request.user.username
            )
            email = request.POST.get("email") or request.user.email
            user = request.user
        else:
            full_name = request.POST.get("full_name")
            email = request.POST.get("email")
            user = None

        phone = request.POST.get("phone")
        address = request.POST.get("address")
        order_action = request.POST.get("order_action")

        if not full_name or not phone or not address:
            messages.error(request, "Veuillez remplir le nom, le téléphone et l'adresse.")
            return redirect("store:checkout")

        for item in cart_items:
            if item.product:
                product = item.product

                if product.stock <= 0:
                    messages.error(
                        request,
                        f"Le produit {product.nom} est en rupture de stock."
                    )
                    return redirect("store:cart")

                if item.quantity > product.stock:
                    messages.error(
                        request,
                        f"La quantité demandée pour {product.nom} dépasse le stock disponible. Stock actuel : {product.stock}"
                    )
                    return redirect("store:cart")

        order = CustomerOrder.objects.create(
            user=user,
            session_key=request.session.session_key,
            full_name=full_name,
            phone=phone,
            email=email,
            address=address,
            total_products=total,
            shipping_cost=shipping_cost,
            total_to_pay=total_with_shipping,
            sent_to_whatsapp=False,
        )

        for item in cart_items:
            if item.product:
                product = item.product
                name = product.nom
                unit_price = item.get_unit_price()

                CustomerOrderItem.objects.create(
                    order=order,
                    product=product,
                    offre=None,
                    name=name,
                    quantity=item.quantity,
                    unit_price=unit_price,
                    total_price=item.get_price(),
                )

            elif item.offre:
                name = item.offre.nom
                unit_price = item.get_unit_price()

                CustomerOrderItem.objects.create(
                    order=order,
                    product=None,
                    offre=item.offre,
                    name=name,
                    quantity=item.quantity,
                    unit_price=unit_price,
                    total_price=item.get_price(),
                )

        generate_order_receipt_image(order)

        request.session["last_order_id"] = order.id

        message = "Bonjour, je souhaite confirmer une commande.\n\n"

        message += "===== COMMANDE =====\n"
        message += f"Commande N° : {order.id}\n"
        message += f"Nom du client : {full_name}\n"
        message += f"Telephone : {phone}\n"
        message += f"Adresse : {address}\n"

        if email:
            message += f"Email : {email}\n"

        message += "\n===== PRODUITS COMMANDÉS =====\n"

        for order_item in order.items.all():
            message += (
                f"- {order_item.name}\n"
                f"  Quantité : {order_item.quantity}\n"
                f"  Prix unitaire : {order_item.unit_price:.2f} MAD\n"
                f"  Total : {order_item.total_price:.2f} MAD\n"
            )

        message += "\n===== TOTAL =====\n"
        message += f"Total produits : {total:.2f} MAD\n"
        message += f"Livraison : {shipping_cost:.2f} MAD\n"
        message += f"Total à payer : {total_with_shipping:.2f} MAD\n"

        if order.receipt_image:
            receipt_url = request.build_absolute_uri(order.receipt_image.url)
            message += f"\nBon de commande : {receipt_url}\n"

        message += "\nMerci."

        cart_obj.store_items.all().delete()

        if order_action == "whatsapp":
            order.sent_to_whatsapp = True
            order.save(update_fields=["sent_to_whatsapp"])

            encoded_message = quote(message)
            phone_whatsapp = "212661440367"
            whatsapp_link = f"https://wa.me/{phone_whatsapp}?text={encoded_message}"
            return redirect(whatsapp_link)

        order.sent_to_whatsapp = False
        order.save(update_fields=["sent_to_whatsapp"])

        messages.success(request, "Votre commande a été enregistrée avec succès.")
        return redirect("store:order_success")

    return render(request, "store/checkout.html", {
        "cart": cart_obj,
        "cart_items": cart_items,
        "total": total,
        "shipping_cost": shipping_cost,
        "total_with_shipping": total_with_shipping,
    })


def order_success(request):
    last_order = None
    last_order_id = request.session.get("last_order_id")

    if last_order_id:
        last_order = CustomerOrder.objects.filter(id=last_order_id).first()

    return render(request, "store/order_success.html", {
        "last_order": last_order,
    })


def passe_order(request):
    return redirect("store:checkout")


# ============================================================
# ORDER RECEIPT IMAGE
# ============================================================

def generate_order_receipt_image(order):
    width = 1000
    padding = 45
    line_height = 34

    items = order.items.all()

    base_height = 540
    item_height = max(items.count(), 1) * 90
    height = base_height + item_height

    image = Image.new("RGB", (width, height), "#ffffff")
    draw = ImageDraw.Draw(image)

    try:
        title_font = ImageFont.truetype("arial.ttf", 38)
        subtitle_font = ImageFont.truetype("arial.ttf", 24)
        normal_font = ImageFont.truetype("arial.ttf", 20)
        small_font = ImageFont.truetype("arial.ttf", 18)
        brand_font = ImageFont.truetype("arial.ttf", 24)
        brand_small_font = ImageFont.truetype("arial.ttf", 16)
    except Exception:
        title_font = ImageFont.load_default()
        subtitle_font = ImageFont.load_default()
        normal_font = ImageFont.load_default()
        small_font = ImageFont.load_default()
        brand_font = ImageFont.load_default()
        brand_small_font = ImageFont.load_default()

    header_color = "#0b777b"
    draw.rectangle((0, 0, width, 140), fill=header_color)
    draw.text((padding, 30), "BON DE COMMANDE", fill="white", font=title_font)
    draw.text((padding, 85), f"Commande N° {order.id}", fill="white", font=subtitle_font)

    brand_box_x1 = width - 330
    brand_box_y1 = 18
    brand_box_x2 = width - 35
    brand_box_y2 = 118

    draw.rounded_rectangle(
        (brand_box_x1, brand_box_y1, brand_box_x2, brand_box_y2),
        radius=18,
        fill="white"
    )

    draw.text(
        (brand_box_x1 + 20, brand_box_y1 + 18),
        "City Pharma Plus",
        fill="#0b777b",
        font=brand_font
    )
    draw.text(
        (brand_box_x1 + 20, brand_box_y1 + 55),
        "VOTRE REUSSITE, NOTRE METIER",
        fill="#4d6f70",
        font=brand_small_font
    )

    y = 180

    draw.text((padding, y), f"Client : {order.full_name}", fill="#222222", font=normal_font)
    y += line_height

    draw.text((padding, y), f"Telephone : {order.phone}", fill="#222222", font=normal_font)
    y += line_height

    if order.email:
        draw.text((padding, y), f"Email : {order.email}", fill="#222222", font=normal_font)
        y += line_height

    address_lines = wrap(f"Adresse : {order.address}", width=70)
    for line in address_lines:
        draw.text((padding, y), line, fill="#222222", font=normal_font)
        y += line_height

    y += 25

    draw.rectangle((padding, y, width - padding, y + 48), fill="#eef3f3")
    draw.text((padding + 15, y + 13), "Produit", fill="#111111", font=small_font)
    draw.text((width - 330, y + 13), "Qté", fill="#111111", font=small_font)
    draw.text((width - 220, y + 13), "Total", fill="#111111", font=small_font)

    y += 65

    for item in items:
        product_lines = wrap(item.name, width=50) or [item.name]

        draw.text((padding + 15, y), product_lines[0], fill="#222222", font=normal_font)
        draw.text((width - 330, y), str(item.quantity), fill="#222222", font=normal_font)
        draw.text((width - 220, y), f"{item.total_price:.2f} DH", fill="#222222", font=normal_font)

        y += 32

        for line in product_lines[1:]:
            draw.text((padding + 15, y), line, fill="#555555", font=small_font)
            y += 28

        draw.line((padding, y + 10, width - padding, y + 10), fill="#e8e8e8", width=1)
        y += 35

    y += 20

    draw.rounded_rectangle(
        (padding, y, width - padding, y + 160),
        radius=18,
        fill="#f7f9f9"
    )

    draw.text(
        (padding + 22, y + 25),
        f"Montant produits : {order.total_products:.2f} DH",
        fill="#222222",
        font=normal_font
    )
    draw.text(
        (padding + 22, y + 70),
        f"Livraison : {order.shipping_cost:.2f} DH",
        fill="#222222",
        font=normal_font
    )
    draw.text(
        (padding + 22, y + 115),
        f"Total à payer : {order.total_to_pay:.2f} DH",
        fill="#0b777b",
        font=subtitle_font
    )

    buffer = BytesIO()
    image.save(buffer, format="PNG")

    file_name = f"commande_{order.id}.png"

    order.receipt_image.save(
        file_name,
        ContentFile(buffer.getvalue()),
        save=True
    )


# ============================================================
# MARQUES
# ============================================================

def produits_par_marque(request, marque_nom):
    marque = get_object_or_404(Marque, nom__iexact=marque_nom)

    produits_queryset = (
        Product.objects
        .filter(marque=marque, active=True)
        .select_related("marque", "categorie", "sous_categorie")
        .order_by("-id")
    )

    produits, pagination_query = paginate_queryset(
        request,
        produits_queryset,
        PRODUCTS_PER_PAGE
    )

    return render(request, "store/produits_par_marque.html", {
        "marque": marque,
        "produits": produits,
        "page_obj": produits,
        "pagination_query": pagination_query,
    })


def marques_et_produits(request):
    marques = (
        Marque.objects
        .prefetch_related(
            Prefetch(
                "products",
                queryset=Product.objects
                .filter(active=True)
                .select_related("categorie", "sous_categorie")
            )
        )
        .order_by("nom")
    )

    return render(request, "store/marques_et_produits.html", {
        "marques": marques,
    })


def detail_marque(request, marque_id):
    marque = get_object_or_404(Marque, id=marque_id)

    produits_queryset = (
        Product.objects
        .filter(marque=marque, active=True)
        .select_related("marque", "categorie", "sous_categorie")
        .order_by("-id")
    )

    sous_categories = []
    seen = set()

    for produit in produits_queryset:
        if produit.sous_categorie_id and produit.sous_categorie_id not in seen:
            sous_categories.append(produit.sous_categorie)
            seen.add(produit.sous_categorie_id)

    produits, pagination_query = paginate_queryset(
        request,
        produits_queryset,
        PRODUCTS_PER_PAGE
    )

    return render(request, "marque/detail_marque.html", {
        "marque": marque,
        "produits": produits,
        "page_obj": produits,
        "pagination_query": pagination_query,
        "sous_categories": sous_categories,
    })


def marques_grid(request):
    marques_queryset = (
        Marque.objects
        .annotate(
            nbr_produits=Count(
                "products",
                filter=Q(products__active=True),
                distinct=True
            )
        )
        .order_by("nom")
    )

    marques, pagination_query = paginate_queryset(
        request,
        marques_queryset,
        BRANDS_PER_PAGE
    )

    return render(request, "marque/liste_marques.html", {
        "marques": marques,
        "page_obj": marques,
        "pagination_query": pagination_query,
    })
# ============================================================
# PAGES
# ============================================================

def qui_sommes_nous(request):
    return render(request, "store/qui.html")


# ============================================================
# OFFRES
# ============================================================

def liste_offres(request):
    offres_queryset = (
        OffreSpeciale.objects
        .filter(
            Q(date_expiration__isnull=True) |
            Q(date_expiration__gt=timezone.now())
        )
        .order_by("-id")
    )

    offres, pagination_query = paginate_queryset(
        request,
        offres_queryset,
        OFFERS_PER_PAGE
    )

    return render(request, "store/liste_offres.html", {
        "offres": offres,
        "page_obj": offres,
        "pagination_query": pagination_query,
    })


def detail_offre(request, slug):
    offre = get_object_or_404(OffreSpeciale, slug=slug)

    offres_similaires = (
        OffreSpeciale.objects
        .filter(
            Q(date_expiration__isnull=True) |
            Q(date_expiration__gt=timezone.now())
        )
        .exclude(id=offre.id)
        .order_by("-id")[:4]
    )

    return render(request, "store/detail_offre.html", {
        "offre": offre,
        "offres_similaires": offres_similaires,
    })


# ============================================================
# SEARCH
# ============================================================

def search_products(request):
    query = request.GET.get("q", "").strip()
    produits_queryset = Product.objects.none()

    if query:
        produits_queryset = (
            Product.objects
            .filter(
                Q(nom__icontains=query) |
                Q(description__icontains=query) |
                Q(marque__nom__icontains=query) |
                Q(sous_categorie__nom__icontains=query),
                active=True
            )
            .select_related("marque", "sous_categorie")
            .distinct()
            .order_by("-id")
        )

    produits, pagination_query = paginate_queryset(
        request,
        produits_queryset,
        PRODUCTS_PER_PAGE
    )

    return render(request, "store/search_results.html", {
        "query": query,
        "produits": produits,
        "page_obj": produits,
        "pagination_query": pagination_query,
        "total_results": produits_queryset.count(),
    })


# ============================================================
# STOCK / SCANNER / DASHBOARD
# ============================================================

@staff_member_required
def scanner_stock(request):
    product = None
    barcode = ""
    message = ""
    add_product_url = ""
    q = request.GET.get("q", "").strip()
    search_results = []

    # Recherche rapide par nom / marque / code-barres
    if q:
        search_results = (
            Product.objects
            .filter(
                Q(nom__icontains=q) |
                Q(barcode__icontains=q) |
                Q(marque__nom__icontains=q)
            )
            .select_related("marque", "categorie")
            .order_by("nom")[:20]
        )

    if request.method == "POST":
        barcode = request.POST.get("barcode", "").strip()
        action = request.POST.get("action", "search")

        try:
            quantity = int(request.POST.get("quantity", 1))
            if quantity < 1:
                quantity = 1
        except ValueError:
            quantity = 1

        if barcode:
            product = Product.objects.filter(barcode=barcode).first()

            if product:
                if action == "add":
                    old_stock = product.stock

                    product.stock += quantity
                    product.save(update_fields=["stock"])

                    new_stock = product.stock

                    StockMovement.objects.create(
                        product=product,
                        movement_type="add",
                        quantity=quantity,
                        old_stock=old_stock,
                        new_stock=new_stock,
                        note=f"Ajout de {quantity} depuis scanner",
                        created_by=request.user if request.user.is_authenticated else None,
                    )

                    message = f"تمت زيادة {quantity} للمنتج: {product.nom}"

                elif action == "remove":
                    if product.stock >= quantity:
                        old_stock = product.stock

                        product.stock -= quantity
                        product.save(update_fields=["stock"])

                        new_stock = product.stock

                        StockMovement.objects.create(
                            product=product,
                            movement_type="remove",
                            quantity=quantity,
                            old_stock=old_stock,
                            new_stock=new_stock,
                            note=f"Retrait de {quantity} depuis scanner",
                            created_by=request.user if request.user.is_authenticated else None,
                        )

                        message = f"تم إنقاص {quantity} من المنتج: {product.nom}"
                    else:
                        message = f"لا يمكن إنقاص {quantity}. الكمية المتوفرة فقط هي {product.stock}."

            else:
                message = "هذا المنتج غير موجود بهذا الكود بار."
                add_product_url = f"/admin/store/product/add/?barcode={barcode}"

        else:
            message = "دخلي أو سكانِي الكود بار."

    return render(request, "store/scanner_stock.html", {
        "product": product,
        "barcode": barcode,
        "message": message,
        "add_product_url": add_product_url,
        "q": q,
        "search_results": search_results,
    })


def get_filtered_stock_products(request):
    q = request.GET.get("q", "").strip()
    marque_id = request.GET.get("marque", "").strip()
    categorie_id = request.GET.get("categorie", "").strip()
    stock_status = request.GET.get("stock_status", "").strip()

    products_qs = (
        Product.objects
        .filter(active=True)
        .select_related("categorie", "sous_categorie", "marque")
        .order_by("nom")
    )

    if q:
        products_qs = products_qs.filter(
            Q(nom__icontains=q) |
            Q(barcode__icontains=q) |
            Q(marque__nom__icontains=q) |
            Q(categorie__nom__icontains=q) |
            Q(sous_categorie__nom__icontains=q)
        )

    if marque_id:
        products_qs = products_qs.filter(marque_id=marque_id)

    if categorie_id:
        products_qs = products_qs.filter(categorie_id=categorie_id)

    if stock_status == "missing_barcode":
        products_qs = products_qs.filter(
            Q(barcode__isnull=True) | Q(barcode="")
        )

    elif stock_status == "missing_purchase_price":
        products_qs = products_qs.filter(prix_achat=0)

    products_list = list(products_qs)

    if stock_status == "low":
        products_list = [
            product for product in products_list
            if product.is_low_stock()
        ]

    elif stock_status == "ok":
        products_list = [
            product for product in products_list
            if not product.is_low_stock()
        ]

    return products_list, {
        "q": q,
        "marque_id": marque_id,
        "categorie_id": categorie_id,
        "stock_status": stock_status,
    }


@staff_member_required
def stock_dashboard(request):
    products_list, filters = get_filtered_stock_products(request)

    total_products = len(products_list)
    total_quantity = sum(product.stock for product in products_list)

    total_stock_achat = sum(
        (product.valeur_stock_achat() for product in products_list),
        Decimal("0.00")
    )

    total_stock_vente = sum(
        (product.valeur_stock_vente() for product in products_list),
        Decimal("0.00")
    )

    total_benefice = total_stock_vente - total_stock_achat

    low_stock_products = [
        product for product in products_list
        if product.is_low_stock()
    ]

    paginator = Paginator(products_list, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    marques = Marque.objects.all().order_by("nom")
    categories = Categorie.objects.filter(active=True).order_by("ordre", "nom")

    return render(request, "store/stock_dashboard.html", {
        "products": page_obj,
        "page_obj": page_obj,

        "total_products": total_products,
        "total_quantity": total_quantity,
        "total_stock_achat": total_stock_achat,
        "total_stock_vente": total_stock_vente,
        "total_benefice": total_benefice,
        "low_stock_products": low_stock_products,

        "marques": marques,
        "categories": categories,

        "q": filters["q"],
        "selected_marque": filters["marque_id"],
        "selected_categorie": filters["categorie_id"],
        "selected_stock_status": filters["stock_status"],
    })


@staff_member_required
def export_stock_excel(request):  
    products, filters = get_filtered_stock_products(request)

    total_products = len(products)
    total_quantity = sum(product.stock for product in products)

    total_stock_achat = sum(
        (product.valeur_stock_achat() for product in products),
        Decimal("0.00")
    )

    total_stock_vente = sum(
        (product.valeur_stock_vente() for product in products),
        Decimal("0.00")
    )

    total_benefice = total_stock_vente - total_stock_achat

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"

    ws["A1"] = "Rapport de Stock"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:M1")

    summary_data = [
        ("Nombre de produits", total_products),
        ("Quantité totale", total_quantity),
        ("Valeur achat", float(total_stock_achat)),
        ("Valeur vente", float(total_stock_vente)),
        ("Bénéfice potentiel", float(total_benefice)),
    ]

    row = 3
    for label, value in summary_data:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

    headers = [
        "Produit",
        "Catégorie",
        "Sous-catégorie",
        "Marque",
        "Code-barres",
        "Prix achat",
        "Prix vente",
        "Stock",
        "Stock minimum",
        "Valeur achat",
        "Valeur vente",
        "Bénéfice potentiel",
        "État",
    ]

    start_row = 10

    header_fill = PatternFill("solid", fgColor="00796B")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_num, product in enumerate(products, start_row + 1):
        prix_vente = product.get_price()
        valeur_achat = product.valeur_stock_achat()
        valeur_vente = product.valeur_stock_vente()
        benefice = product.benefice_potentiel()
        etat = "Stock faible" if product.is_low_stock() else "OK"

        ws.cell(row=row_num, column=1, value=product.nom)
        ws.cell(row=row_num, column=2, value=product.categorie.nom if product.categorie else "")
        ws.cell(row=row_num, column=3, value=product.sous_categorie.nom if product.sous_categorie else "")
        ws.cell(row=row_num, column=4, value=product.marque.nom if product.marque else "")
        ws.cell(row=row_num, column=5, value=product.barcode or "")
        ws.cell(row=row_num, column=6, value=float(product.prix_achat))
        ws.cell(row=row_num, column=7, value=float(prix_vente))
        ws.cell(row=row_num, column=8, value=product.stock)
        ws.cell(row=row_num, column=9, value=product.stock_min)
        ws.cell(row=row_num, column=10, value=float(valeur_achat))
        ws.cell(row=row_num, column=11, value=float(valeur_vente))
        ws.cell(row=row_num, column=12, value=float(benefice))
        ws.cell(row=row_num, column=13, value=etat)

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="rapport_stock.xlsx"'

    wb.save(response)
    return response


@staff_member_required
def stock_movements(request):
    movements = (
        StockMovement.objects
        .select_related("product", "product__marque", "created_by")
        .order_by("-created_at")[:200]
    )

    return render(request, "store/stock_movements.html", {
        "movements": movements,
    })